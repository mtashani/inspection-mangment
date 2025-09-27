# Reporting Service for Data Export and Report Generation
from typing import Dict, List, Optional, Any
from sqlmodel import Session, select, text
from datetime import datetime, date, timedelta
from io import BytesIO, StringIO
import csv
import json
from app.domains.inspector.models.inspector import Inspector
from app.domains.inspector.models.attendance import AttendanceRecord, MonthlyAttendance
from app.domains.inspector.models.enums import AttendanceStatus
from app.common.utils import jalali_calendar


class ReportingService:
    def __init__(self, db: Session):
        self.db = db

    def generate_attendance_report(self, filters: Dict) -> Dict:
        """Generate comprehensive attendance report based on filters"""
        
        # Parse filters
        start_date = filters.get('start_date')
        end_date = filters.get('end_date')
        inspector_ids = filters.get('inspector_ids', [])
        departments = filters.get('departments', [])
        include_details = filters.get('include_details', True)
        report_type = filters.get('report_type', 'summary')  # summary, detailed, analytics
        
        # Convert Jalali dates if provided
        if filters.get('jalali_start_date'):
            year, month, day = map(int, filters['jalali_start_date'].split('-'))
            start_date = jalali_calendar.jalali_to_gregorian(year, month, day)
        
        if filters.get('jalali_end_date'):
            year, month, day = map(int, filters['jalali_end_date'].split('-'))
            end_date = jalali_calendar.jalali_to_gregorian(year, month, day)
        
        # Default to current month if no dates provided
        if not start_date or not end_date:
            today = date.today()
            jalali_today = jalali_calendar.gregorian_to_jalali_str(today)
            jalali_year, jalali_month, _ = map(int, jalali_today.split('-'))
            start_date = jalali_calendar.jalali_to_gregorian(jalali_year, jalali_month, 1)
            days_in_month = jalali_calendar.get_jalali_month_days(jalali_year, jalali_month)
            end_date = jalali_calendar.jalali_to_gregorian(jalali_year, jalali_month, days_in_month)
        
        # Build query
        query = select(AttendanceRecord).where(
            AttendanceRecord.date >= start_date,
            AttendanceRecord.date <= end_date
        )
        
        if inspector_ids:
            query = query.where(AttendanceRecord.inspector_id.in_(inspector_ids))
        
        records = self.db.exec(query).all()
        
        # Get inspector details
        inspector_query = select(Inspector).where(Inspector.active == True)
        if inspector_ids:
            inspector_query = inspector_query.where(Inspector.id.in_(inspector_ids))
        if departments:
            inspector_query = inspector_query.where(Inspector.department.in_(departments))
        
        inspectors = self.db.exec(inspector_query).all()
        inspector_dict = {inspector.id: inspector for inspector in inspectors}
        
        # Generate report based on type
        if report_type == 'summary':
            report_data = self._generate_summary_report(records, inspector_dict, start_date, end_date)
        elif report_type == 'detailed':
            report_data = self._generate_detailed_report(records, inspector_dict, start_date, end_date)
        else:  # analytics
            report_data = self._generate_analytics_report(records, inspector_dict, start_date, end_date)
        
        return {
            "report_id": f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            "generated_at": datetime.utcnow().isoformat(),
            "filters": filters,
            "period": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "jalali_start_date": jalali_calendar.gregorian_to_jalali_str(start_date),
                "jalali_end_date": jalali_calendar.gregorian_to_jalali_str(end_date)
            },
            "report_type": report_type,
            "data": report_data
        }

    def export_to_csv(self, report_data: Dict) -> str:
        """Export report data to CSV format"""
        
        output = StringIO()
        
        report_type = report_data.get('report_type', 'summary')
        data = report_data.get('data', {})
        
        if report_type == 'summary':
            # Export summary data
            writer = csv.writer(output)
            
            # Header
            writer.writerow([
                'Inspector ID', 'Inspector Name', 'Employee ID', 'Department',
                'Working Days', 'Resting Days', 'Leave Days', 'Absent Days',
                'Total Regular Hours', 'Total Overtime Hours', 'Attendance Rate'
            ])
            
            # Data rows
            for inspector_summary in data.get('inspector_summaries', []):
                writer.writerow([
                    inspector_summary.get('inspector_id'),
                    inspector_summary.get('inspector_name'),
                    inspector_summary.get('employee_id'),
                    inspector_summary.get('department'),
                    inspector_summary.get('working_days'),
                    inspector_summary.get('resting_days'),
                    inspector_summary.get('leave_days'),
                    inspector_summary.get('absent_days'),
                    inspector_summary.get('total_regular_hours'),
                    inspector_summary.get('total_overtime_hours'),
                    inspector_summary.get('attendance_rate')
                ])
        
        elif report_type == 'detailed':
            # Export detailed attendance records
            writer = csv.writer(output)
            
            # Header
            writer.writerow([
                'Date', 'Jalali Date', 'Inspector ID', 'Inspector Name', 
                'Employee ID', 'Department', 'Status', 'Regular Hours', 
                'Overtime Hours', 'Night Shift Hours', 'On Call Hours', 
                'Is Override', 'Override Reason', 'Notes'
            ])
            
            # Data rows
            for record in data.get('detailed_records', []):
                writer.writerow([
                    record.get('date'),
                    record.get('jalali_date'),
                    record.get('inspector_id'),
                    record.get('inspector_name'),
                    record.get('employee_id'),
                    record.get('department'),
                    record.get('status'),
                    record.get('regular_hours'),
                    record.get('overtime_hours'),
                    record.get('night_shift_hours'),
                    record.get('on_call_hours'),
                    record.get('is_override'),
                    record.get('override_reason'),
                    record.get('notes')
                ])
        
        csv_content = output.getvalue()
        output.close()
        
        return csv_content

    def export_to_excel(self, report_data: Dict) -> bytes:
        """Export report data to Excel format"""
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ValueError("openpyxl library is required for Excel export")
        
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        report_type = report_data.get('report_type', 'summary')
        data = report_data.get('data', {})
        
        if report_type == 'summary':
            self._create_summary_excel_sheet(workbook, data, report_data)
        elif report_type == 'detailed':
            self._create_detailed_excel_sheet(workbook, data, report_data)
        else:  # analytics
            self._create_analytics_excel_sheet(workbook, data, report_data)
        
        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output.read()

    def export_to_pdf(self, report_data: Dict) -> bytes:
        """Export report data to PDF format"""
        
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
        except ImportError:
            raise ValueError("reportlab library is required for PDF export")
        
        output = BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=1,  # Center alignment
            spaceAfter=30
        )
        
        title = Paragraph("Attendance Report", title_style)
        story.append(title)
        
        # Report info
        period = report_data.get('period', {})
        info_text = f"""
        <b>Report Type:</b> {report_data.get('report_type', 'N/A').title()}<br/>
        <b>Period:</b> {period.get('jalali_start_date', 'N/A')} to {period.get('jalali_end_date', 'N/A')}<br/>
        <b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        
        info_para = Paragraph(info_text, styles['Normal'])
        story.append(info_para)
        story.append(Spacer(1, 20))
        
        # Data table
        data_for_table = self._prepare_pdf_table_data(report_data)
        if data_for_table:
            table = Table(data_for_table)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
        
        doc.build(story)
        output.seek(0)
        
        return output.read()

    def get_export_formats(self) -> List[str]:
        """Get list of supported export formats"""
        return ['csv', 'excel', 'pdf', 'json']

    def bulk_export_data(self, export_format: str, filters: Dict) -> bytes:
        """Bulk export data in specified format"""
        
        # Generate report
        report_data = self.generate_attendance_report(filters)
        
        if export_format.lower() == 'csv':
            content = self.export_to_csv(report_data)
            return content.encode('utf-8')
        elif export_format.lower() == 'excel':
            return self.export_to_excel(report_data)
        elif export_format.lower() == 'pdf':
            return self.export_to_pdf(report_data)
        elif export_format.lower() == 'json':
            content = json.dumps(report_data, indent=2, ensure_ascii=False, default=str)
            return content.encode('utf-8')
        else:
            raise ValueError(f"Unsupported export format: {export_format}")

    def _generate_summary_report(self, records: List[AttendanceRecord], inspector_dict: Dict, start_date: date, end_date: date) -> Dict:
        """Generate summary report data"""
        
        # Group records by inspector
        records_by_inspector = {}
        for record in records:
            if record.inspector_id not in records_by_inspector:
                records_by_inspector[record.inspector_id] = []
            records_by_inspector[record.inspector_id].append(record)
        
        inspector_summaries = []
        total_stats = {
            'total_working_days': 0,
            'total_resting_days': 0,
            'total_leave_days': 0,
            'total_absent_days': 0,
            'total_regular_hours': 0,
            'total_overtime_hours': 0
        }
        
        for inspector_id, inspector_records in records_by_inspector.items():
            inspector = inspector_dict.get(inspector_id)
            if not inspector:
                continue
            
            working_days = len([r for r in inspector_records if r.status == AttendanceStatus.WORKING])
            resting_days = len([r for r in inspector_records if r.status == AttendanceStatus.RESTING])
            leave_days = len([r for r in inspector_records if r.status == AttendanceStatus.LEAVE])
            absent_days = len([r for r in inspector_records if r.status == AttendanceStatus.ABSENT])
            
            total_regular_hours = sum(r.regular_hours for r in inspector_records)
            total_overtime_hours = sum(r.overtime_hours for r in inspector_records)
            
            attendance_rate = (working_days / max(len(inspector_records), 1)) * 100
            
            summary = {
                'inspector_id': inspector_id,
                'inspector_name': f"{inspector.first_name} {inspector.last_name}",
                'employee_id': inspector.employee_id,
                'working_days': working_days,
                'resting_days': resting_days,
                'leave_days': leave_days,
                'absent_days': absent_days,
                'total_days': len(inspector_records),
                'total_regular_hours': total_regular_hours,
                'total_overtime_hours': total_overtime_hours,
                'attendance_rate': round(attendance_rate, 1)
            }
            
            inspector_summaries.append(summary)
            
            # Update totals
            total_stats['total_working_days'] += working_days
            total_stats['total_resting_days'] += resting_days
            total_stats['total_leave_days'] += leave_days
            total_stats['total_absent_days'] += absent_days
            total_stats['total_regular_hours'] += total_regular_hours
            total_stats['total_overtime_hours'] += total_overtime_hours
        
        return {
            'inspector_summaries': inspector_summaries,
            'total_stats': total_stats,
            'period_info': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'total_days': (end_date - start_date).days + 1
            }
        }

    def _generate_detailed_report(self, records: List[AttendanceRecord], inspector_dict: Dict, start_date: date, end_date: date) -> Dict:
        """Generate detailed report data"""
        
        detailed_records = []
        
        for record in records:
            inspector = inspector_dict.get(record.inspector_id)
            if not inspector:
                continue
            
            detailed_record = {
                'date': record.date.isoformat(),
                'jalali_date': jalali_calendar.gregorian_to_jalali_str(record.date),
                'inspector_id': record.inspector_id,
                'inspector_name': f"{inspector.first_name} {inspector.last_name}",
                'employee_id': inspector.employee_id,
                'status': record.status.value,
                'regular_hours': record.regular_hours,
                'overtime_hours': record.overtime_hours,
                'night_shift_hours': record.night_shift_hours,
                'on_call_hours': record.on_call_hours,
                'is_override': record.is_override,
                'override_reason': record.override_reason,
                'notes': record.notes,
                'created_at': record.created_at.isoformat(),
                'updated_at': record.updated_at.isoformat()
            }
            
            detailed_records.append(detailed_record)
        
        # Sort by date and inspector
        detailed_records.sort(key=lambda x: (x['date'], x['inspector_name']))
        
        return {
            'detailed_records': detailed_records,
            'total_records': len(detailed_records),
            'period_info': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'total_days': (end_date - start_date).days + 1
            }
        }

    def _generate_analytics_report(self, records: List[AttendanceRecord], inspector_dict: Dict, start_date: date, end_date: date) -> Dict:
        """Generate analytics report data"""
        
        # Import analytics service for calculations
        from app.domains.inspector.services.analytics_service import AnalyticsService
        analytics_service = AnalyticsService(self.db)
        
        # Get various analytics
        overview = analytics_service.get_attendance_overview("custom")
        
        # Department-wise analysis
        dept_analysis = {}
        for record in records:
            inspector = inspector_dict.get(record.inspector_id)
            if not inspector:
                continue
            
            dept = 'Unknown'
            if dept not in dept_analysis:
                dept_analysis[dept] = {
                    'working_days': 0,
                    'total_days': 0,
                    'total_overtime': 0,
                    'inspectors': set()
                }
            
            dept_analysis[dept]['total_days'] += 1
            dept_analysis[dept]['total_overtime'] += record.overtime_hours
            dept_analysis[dept]['inspectors'].add(record.inspector_id)
            
            if record.status == AttendanceStatus.WORKING:
                dept_analysis[dept]['working_days'] += 1
        
        # Convert sets to counts
        for dept_data in dept_analysis.values():
            dept_data['inspectors'] = len(dept_data['inspectors'])
            dept_data['attendance_rate'] = (dept_data['working_days'] / max(dept_data['total_days'], 1)) * 100
        
        return {
            'overview': overview,
            'department_analysis': dept_analysis,
            'period_info': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'total_days': (end_date - start_date).days + 1
            }
        }

    def _create_summary_excel_sheet(self, workbook, data: Dict, report_data: Dict):
        """Create summary sheet in Excel workbook"""
        
        sheet = workbook.create_sheet("Summary Report")
        
        # Headers
        headers = [
            'Inspector ID', 'Inspector Name', 'Employee ID', 'Department',
            'Working Days', 'Resting Days', 'Leave Days', 'Absent Days',
            'Total Regular Hours', 'Total Overtime Hours', 'Attendance Rate'
        ]
        
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Data
        for row, summary in enumerate(data.get('inspector_summaries', []), 2):
            sheet.cell(row=row, column=1, value=summary.get('inspector_id'))
            sheet.cell(row=row, column=2, value=summary.get('inspector_name'))
            sheet.cell(row=row, column=3, value=summary.get('employee_id'))
            sheet.cell(row=row, column=4, value=summary.get('department'))
            sheet.cell(row=row, column=5, value=summary.get('working_days'))
            sheet.cell(row=row, column=6, value=summary.get('resting_days'))
            sheet.cell(row=row, column=7, value=summary.get('leave_days'))
            sheet.cell(row=row, column=8, value=summary.get('absent_days'))
            sheet.cell(row=row, column=9, value=summary.get('total_regular_hours'))
            sheet.cell(row=row, column=10, value=summary.get('total_overtime_hours'))
            sheet.cell(row=row, column=11, value=summary.get('attendance_rate'))

    def _create_detailed_excel_sheet(self, workbook, data: Dict, report_data: Dict):
        """Create detailed sheet in Excel workbook"""
        
        sheet = workbook.create_sheet("Detailed Report")
        
        # Headers
        headers = [
            'Date', 'Jalali Date', 'Inspector Name', 'Employee ID', 
            'Department', 'Status', 'Regular Hours', 'Overtime Hours', 
            'Is Override', 'Notes'
        ]
        
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Data
        for row, record in enumerate(data.get('detailed_records', []), 2):
            sheet.cell(row=row, column=1, value=record.get('date'))
            sheet.cell(row=row, column=2, value=record.get('jalali_date'))
            sheet.cell(row=row, column=3, value=record.get('inspector_name'))
            sheet.cell(row=row, column=4, value=record.get('employee_id'))
            sheet.cell(row=row, column=5, value=record.get('department'))
            sheet.cell(row=row, column=6, value=record.get('status'))
            sheet.cell(row=row, column=7, value=record.get('regular_hours'))
            sheet.cell(row=row, column=8, value=record.get('overtime_hours'))
            sheet.cell(row=row, column=9, value=record.get('is_override'))
            sheet.cell(row=row, column=10, value=record.get('notes'))

    def _create_analytics_excel_sheet(self, workbook, data: Dict, report_data: Dict):
        """Create analytics sheet in Excel workbook"""
        
        sheet = workbook.create_sheet("Analytics Report")
        
        # Add overview data
        overview = data.get('overview', {})
        sheet.cell(row=1, column=1, value="Overall Statistics")
        sheet.cell(row=2, column=1, value="Total Records")
        sheet.cell(row=2, column=2, value=overview.get('total_records'))
        sheet.cell(row=3, column=1, value="Attendance Rate")
        sheet.cell(row=3, column=2, value=overview.get('attendance_rate'))
        
        # Add department analysis
        dept_analysis = data.get('department_analysis', {})
        if dept_analysis:
            sheet.cell(row=5, column=1, value="Department Analysis")
            headers = ['Department', 'Inspectors', 'Working Days', 'Total Days', 'Attendance Rate', 'Total Overtime']
            
            for col, header in enumerate(headers, 1):
                sheet.cell(row=6, column=col, value=header)
            
            for row, (dept, dept_data) in enumerate(dept_analysis.items(), 7):
                sheet.cell(row=row, column=1, value=dept)
                sheet.cell(row=row, column=2, value=dept_data.get('inspectors'))
                sheet.cell(row=row, column=3, value=dept_data.get('working_days'))
                sheet.cell(row=row, column=4, value=dept_data.get('total_days'))
                sheet.cell(row=row, column=5, value=dept_data.get('attendance_rate'))
                sheet.cell(row=row, column=6, value=dept_data.get('total_overtime'))

    def _prepare_pdf_table_data(self, report_data: Dict) -> List[List[str]]:
        """Prepare data for PDF table format"""
        
        report_type = report_data.get('report_type', 'summary')
        data = report_data.get('data', {})
        
        if report_type == 'summary':
            table_data = [
                ['Inspector', 'Employee ID', 'Working Days', 'Leave Days', 'Rate %', 'Overtime']
            ]
            
            for summary in data.get('inspector_summaries', []):
                table_data.append([
                    summary.get('inspector_name', '')[:20],  # Truncate for PDF
                    summary.get('employee_id', ''),
                    str(summary.get('working_days', 0)),
                    str(summary.get('leave_days', 0)),
                    f"{summary.get('attendance_rate', 0):.1f}%",
                    f"{summary.get('total_overtime_hours', 0):.1f}h"
                ])
            
            return table_data
        
        return []